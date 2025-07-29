import pandas as pd
import glob
from fpdf import FPDF, XPos, YPos
from pathlib import Path

from openpyxl.styles.builtins import styles

filepaths = glob.glob("invoces/*.xlsx");

for filepath in filepaths:
    # Extract invoces data
    pdf = FPDF(orientation="P", unit="mm", format="A4");
    pdf.add_page();

    # Get the correct path name
    filename = Path(filepath).stem;
    invoice_nm = filename.split("-")[0];
    date = filename.split("-")[1];

    #Creating pdf
    pdf.set_font(family="Times", size=16, style="B");
    pdf.cell(w=50, h=8, text=f"invoice nr. {invoice_nm}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.set_font(family="Times", size=16, style="B");
    pdf.cell(w=50, h=8, text=f"Date .{date}", new_x=XPos.LMARGIN, new_y=YPos.NEXT);

    # Extract data
    df = pd.read_excel(filepath, sheet_name="Sheet 1");
    columns = list(df.columns);
    # Remove "-"
    columns = [item.replace("-"," ").title() for item in columns];

    # Adding header
    pdf.set_font(family="Times", size=10, style="B");
    pdf.set_text_color(80, 80, 80);
    pdf.cell(w=30, h=8, text=columns[0], border=1);
    pdf.cell(w=70, h=8, text=columns[1], border=1);
    pdf.cell(w=30, h=8, text=columns[2], border=1);
    pdf.cell(w=30, h=8, text=columns[3], border=1);
    pdf.cell(w=30, h=8, text=columns[4], border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT);

    # Adding body data
    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10);
        pdf.set_text_color(80, 80, 80);
        pdf.cell(w=30, h=8, text=str(row["product_id"]), border=1);
        pdf.cell(w=70, h=8, text=str(row["product_name"]), border=1);
        pdf.cell(w=30, h=8, text=str(row["amount_purchased"]), border=1);
        pdf.cell(w=30, h=8, text=str(row["price_per_unit"]), border=1);
        pdf.cell(w=30, h=8, text=str(row["total_price"]), border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT);

    # Total sum
    total_sum = df["total_price"].sum();
    pdf.set_font(family="Times", size=10);
    pdf.set_text_color(80, 80, 80);
    pdf.cell(w=30, h=8, text="", border=1);
    pdf.cell(w=70, h=8, text="", border=1);
    pdf.cell(w=30, h=8, text="", border=1);
    pdf.cell(w=30, h=8, text="", border=1);
    pdf.cell(w=30, h=8, text=str(total_sum), border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT);

    # Add total sum sentense
    pdf.set_font(family="Times", size=10);
    pdf.cell(w=30, h=8, text=f"The total price is {total_sum}", new_x=XPos.LMARGIN, new_y=YPos.NEXT);

    # Company name and logo
    pdf.set_font(family="Times", size=14, style="B");
    pdf.cell(w=25, h=8, text=f"PythonHow ", new_x=XPos.LMARGIN, new_y=YPos.NEXT);
    pdf.image("pythonhow.png", w=10);

    pdf.output(f"PDFS/{filename}.pdf");